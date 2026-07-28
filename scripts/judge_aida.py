"""LLM-judge grading for an AIDABench run, matching the official evaluator.

AIDABench's own evaluator is an LLM judge: FileEvaluatorAgent.evaluate() takes
(question, reference_file, generated_file) and returns {is_correct, reason},
scored binary per task and averaged. It does NOT use eval_area.

Cell-exact grading is not comparable to that, and on the _en split it is
actively misleading -- source documents carry Chinese-to-English translation
artifacts (a cell whose runs are ['Unit','Unit','1'] reads as "UnitUnit1" for
any correct parser, while the reference says "Unit 1"), so a correct answer
scores as wrong.

This grades an existing run without re-running it, so you can re-judge with a
different model for free.

    OPENROUTER_API_KEY=... python _local/scripts/judge_aida.py --run aida-m3-v2
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import urllib.request
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
BENCH = ROOT / "_local" / "bench" / "aida"
SPLIT = BENCH / "data" / "file_generation_en"

MAX_ROWS = 60
MAX_COLS = 25


def render(path: Path, sheet: str | None = None) -> str:
    """Render a tabular file as compact text for the judge."""

    suffix = path.suffix.lower()
    try:
        if suffix in (".csv", ".txt", ".tsv"):
            import csv

            delim = "\t" if suffix == ".tsv" else ","
            with path.open(encoding="utf-8-sig", newline="", errors="replace") as fh:
                grids = {"(csv)": [r for r in csv.reader(fh, delimiter=delim)]}
        elif suffix == ".xls":
            import pandas as pd

            book = pd.read_excel(path, sheet_name=None, header=None, dtype=object)
            grids = {
                k: v.where(v.notna(), None).values.tolist() for k, v in book.items()
            }
        elif suffix in (".docx", ".doc"):
            import zipfile
            import xml.etree.ElementTree as ET

            W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
            with zipfile.ZipFile(path) as z:
                root = ET.fromstring(z.read("word/document.xml"))
            grids = {}
            for i, tbl in enumerate(root.iter(f"{W}tbl")):
                rows = []
                for tr in tbl.findall(f"{W}tr"):
                    rows.append(
                        [
                            "".join(t.text or "" for t in tc.iter(f"{W}t")).strip()
                            for tc in tr.findall(f"{W}tc")
                        ]
                    )
                grids[f"table{i}"] = rows
        else:
            import openpyxl

            wb = openpyxl.load_workbook(path, data_only=True)
            grids = {
                ws.title: [list(r) for r in ws.iter_rows(values_only=True)]
                for ws in wb.worksheets
            }
    except Exception as exc:
        return f"<<could not read {path.name}: {type(exc).__name__}: {exc}>>"

    out = []
    for name, grid in grids.items():
        nrows = len(grid)
        ncols = max((len(r) for r in grid), default=0)
        out.append(f"### {name}  ({nrows} rows x {ncols} cols)")
        shown = grid[:MAX_ROWS]
        for row in shown:
            cells = ["" if c is None else str(c) for c in row[:MAX_COLS]]
            out.append(" | ".join(c[:40] for c in cells))
        if nrows > MAX_ROWS:
            out.append(f"... ({nrows - MAX_ROWS} more rows omitted)")
    return "\n".join(out)


PROMPT = """\
You are grading whether a generated file correctly fulfils a data task.

## The task the model was given

{question}

## Reference (correct) file: {ref_name}

{ref}

## Generated file: {gen_name}

{gen}

## Measured differences over the whole file

The previews above are truncated. These shapes are computed over the whole
file, so trust them over your reading of the previews:

{facts}

## How to judge

Answer whether the generated file accomplishes what the task asked, using the
reference as the standard for the intended result.

- Judge the DATA, not cosmetics. Column ordering, sheet naming, number
  formatting, and cell styling do not matter unless the task asked for them.
- Small textual differences that trace to source-data quirks are acceptable
  when the underlying records match. This dataset is translated, so source text
  is sometimes mangled (a cell may read "UnitUnit1" where the reference says
  "Unit 1"). Do not fail a result for reproducing its own source faithfully.
- Do fail it for wrong values, wrong row counts, missing or extra records,
  values placed under the wrong column, or the wrong rows selected.
- The row and column counts above cover the whole file. A row-count mismatch
  means records are missing or extra even if the visible rows look right, and
  that is a failure. Do not pass a file on the preview alone when the shapes
  disagree.
- Reordered columns or rows are usually still correct, provided the task did
  not specify an order.

Reply with a single JSON object and nothing else:
{{"is_correct": true or false, "reason": "<one sentence>"}}
"""


def diff_facts(row: dict, produced: Path, ref: Path) -> str:
    """Whole-file comparison stats, so a truncated preview cannot mislead the judge."""

    # Deliberately shape only, no cell match rate. Handing the judge a "42% of
    # cells match" figure makes it echo the strict grader instead of judging --
    # measured on an 18-task sample, agreement jumped from 77% to 94% and the
    # judge stopped rescuing anything. Row and column counts are the facts a
    # truncated preview genuinely hides.
    lines: list[str] = []
    try:
        import openpyxl

        def shape(p: Path):
            if p.suffix.lower() in (".csv", ".txt", ".tsv"):
                with p.open(encoding="utf-8-sig", errors="replace") as fh:
                    rows = fh.read().splitlines()
                return len(rows), (max((r.count(",") + 1 for r in rows[:200]), default=0)), ["(csv)"]
            wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
            ws = wb.worksheets[0]
            return ws.max_row or 0, ws.max_column or 0, wb.sheetnames

        pr, pc, pn = shape(produced)
        rr, rc, rn = shape(ref)
        lines.append(f"- Generated: {pr} rows x {pc} cols, sheets {pn}")
        lines.append(f"- Reference: {rr} rows x {rc} cols, sheets {rn}")
        if pr != rr:
            lines.append(f"- ROW COUNT DIFFERS by {abs(pr-rr)} -- records are missing or extra.")
        if pc != rc:
            lines.append(f"- COLUMN COUNT DIFFERS by {abs(pc-rc)}.")
    except Exception as exc:
        lines.append(f"- Shape comparison unavailable ({type(exc).__name__}).")
    return "\n".join(lines)


def call(prompt: str, model: str, key: str) -> str:
    body = json.dumps(
        {"model": model, "messages": [{"role": "user", "content": prompt}], "temperature": 0}
    ).encode()
    req = urllib.request.Request(
        "https://openrouter.ai/api/v1/chat/completions",
        data=body,
        headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
    )
    with urllib.request.urlopen(req, timeout=300) as r:
        return json.load(r)["choices"][0]["message"]["content"]


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--run", required=True)
    ap.add_argument("--model", default="anthropic/claude-sonnet-5")
    ap.add_argument("--workers", type=int, default=6)
    args = ap.parse_args()

    key = os.environ.get("OPENROUTER_API_KEY")
    if not key:
        print("OPENROUTER_API_KEY not set", file=sys.stderr)
        return 2

    run_root = BENCH / "runs" / args.run
    rows = [json.loads(l) for l in (run_root / "results.jsonl").open(encoding="utf-8")]
    meta = {
        str(r["id"]): r
        for r in (json.loads(l) for l in (SPLIT / "file_generation_en.jsonl").open(encoding="utf-8"))
    }
    print(f"judging {len(rows)} tasks from {args.run} with {args.model}")

    def one(row):
        qid = row["id"]
        rec = meta[qid]
        produced = run_root / "work" / qid / rec["output_file"]
        ref = SPLIT / "reference" / qid / rec["reference_file"]
        if not produced.is_file():
            return {**row, "judge": False, "judge_reason": "output file not created"}
        prompt = PROMPT.format(
            question=rec["question"].strip(),
            ref_name=rec["reference_file"],
            ref=render(ref),
            gen_name=rec["output_file"],
            gen=render(produced),
            facts=diff_facts(row, produced, ref),
        )
        try:
            text = call(prompt, args.model, key)
        except Exception as exc:
            return {**row, "judge": None, "judge_reason": f"judge error: {type(exc).__name__}"}
        m = re.search(r"\{.*\}", text, re.S)
        if not m:
            return {**row, "judge": None, "judge_reason": "unparseable judge output"}
        try:
            verdict = json.loads(m.group(0))
        except json.JSONDecodeError:
            return {**row, "judge": None, "judge_reason": "unparseable judge JSON"}
        return {
            **row,
            "judge": bool(verdict.get("is_correct")),
            "judge_reason": str(verdict.get("reason", ""))[:300],
        }

    out = []
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        for i, r in enumerate(ex.map(one, rows), 1):
            out.append(r)
            print(f"  {i}/{len(rows)}", end="\r", flush=True)
    print()

    path = run_root / "results_judged.jsonl"
    with path.open("w", encoding="utf-8") as fh:
        for r in out:
            fh.write(json.dumps(r) + "\n")

    judged = [r for r in out if r["judge"] is not None]
    errs = len(out) - len(judged)
    n_judge = sum(1 for r in judged if r["judge"])
    n_exact = sum(1 for r in out if r["passed"])
    print("\n" + "=" * 58)
    print(f"tasks                : {len(out)}   ({errs} judge errors excluded)")
    print(f"LLM-judged (official-style) : {n_judge}/{len(judged)} = {100*n_judge/max(len(judged),1):.1f}%")
    print(f"cell-exact (strict, ours)   : {n_exact}/{len(out)} = {100*n_exact/max(len(out),1):.1f}%")
    print(f"\npaper's best pass@1 across all 3 categories: 59.43%")
    agree = sum(1 for r in judged if r["judge"] == r["passed"])
    print(f"\ngraders agree on {agree}/{len(judged)} = {100*agree/max(len(judged),1):.0f}% of tasks")
    print(f"  judged correct but cell-exact failed: {sum(1 for r in judged if r['judge'] and not r['passed'])}")
    print(f"  cell-exact passed but judged wrong  : {sum(1 for r in judged if not r['judge'] and r['passed'])}")
    print(f"\n-> {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
