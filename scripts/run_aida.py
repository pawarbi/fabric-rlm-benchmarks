"""Run fabric-rlm against AIDABench's file_generation_en split.

Differs from SpreadsheetBench in two ways that matter:
  * 41% of tasks are multi-file (up to 13 inputs), so the model must join across
    workbooks rather than edit one in place.
  * The model writes a NEW file whose name the task specifies, instead of
    modifying the input.

Grading here is cell-exact over the task's ``eval_area``, which is deterministic
and free. The official AIDABench evaluator is LLM-based ("coarse-to-fine
structural & content validation"), so scores from this script are NOT directly
comparable to the paper's 59.43% -- cell-exact is the stricter of the two.

    OPENROUTER_API_KEY=... python _local/scripts/run_aida.py --limit 20
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
import time
import traceback
from pathlib import Path

import dspy

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

# This dataset is translated from Chinese and file names, sheet names, and error
# text still carry CJK characters. On Windows the default console codec is
# cp1252, so a single progress line can kill a run that is an hour in -- which
# it did, at task 103 of 240. Never let output encoding end a benchmark.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

from fabric_rlm import File, RLM, SkillLoader  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from aida_trace import save_trace  # noqa: E402

BENCH = ROOT / "_local" / "bench" / "aida"
SPLIT = BENCH / "data" / "file_generation_en"


def split_files(raw: str) -> list[str]:
    """Split a packed filename field into names.

    ``input_file``, ``output_file``, and ``reference_file`` all pack multiple
    names on separate lines (or comma-separated). 13 of 261 tasks emit several
    files, up to 32 in one task. Treating the field as a single filename makes
    the reference unfindable and silently excludes those tasks.
    """

    return [p.strip() for p in re.split(r"[\n,]+", str(raw)) if p.strip()]


# Back-compat alias; the field is not input-specific.
split_inputs = split_files


def parse_eval_area(area: str) -> list[tuple[str | None, str | None, tuple[int, int, int, int]]]:
    """Parse AIDABench's ``eval_area`` into (file, sheet, bounds) targets.

    Only 60% of tasks use a bare ``A2:K854``. The rest qualify by sheet, by
    file, or list several ranges. Observed shapes:

        A2:K854                                     bare
        Sheet1!A1:B2  /  'My Sheet'!A1:B2           sheet-qualified
        sheetMy Sheet!A1:B2                         'sheet' prefix
        ['S'!A1:B2,'T'!C1:D2]                       bracketed list
        [f.xlsx]#['Sheet'!I1:I2]                    file-qualified
        multiple of the above, newline separated

    Grading only the first worksheet, as a naive parser does, silently compares
    the wrong sheet and reports a real answer as wrong.
    """

    from openpyxl.utils import range_boundaries

    if area is None:
        return []
    text = str(area).strip()
    # Upstream typo: "'Sheet1'B2:L90!" puts the separator at the end.
    text = re.sub(r"'([^']+)'([A-Z]+\d+:[A-Z]+\d+)!", r"'\1'!\2", text)
    # "B2: B6" -> "B2:B6"; some rows pad around the colon.
    text = re.sub(r"([A-Z]+\d+)\s*:\s*([A-Z]+\d+)", r"\1:\2", text)
    targets: list[tuple[str | None, str | None, tuple[int, int, int, int]]] = []

    # Split into segments: newlines and semicolons separate independent
    # targets, and a "[file]#[...]" prefix scopes what follows it.
    for chunk in (c.strip() for c in re.split(r"[\n;]+", text) if c.strip()):
        file_hint = None
        m = re.match(r"^\[([^\]]+)\]\s*#\s*(.*)$", chunk)
        if m:
            file_hint, chunk = m.group(1).strip(), m.group(2).strip()
        chunk = chunk.strip()
        if chunk.startswith("[") and chunk.endswith("]"):
            chunk = chunk[1:-1]
        # Split on commas that are not inside quotes.
        for part in re.split(r",(?=(?:[^']*'[^']*')*[^']*$)", chunk):
            part = part.strip().rstrip(",").strip()
            if not part:
                continue
            sheet = None
            if "!" in part:
                sheet, part = part.rsplit("!", 1)
                sheet = sheet.strip().strip("'\"").strip()
                # some rows write "sheetMy Sheet!A1:B2"
                if sheet.startswith("sheet") and len(sheet) > 5 and sheet[5].isupper():
                    sheet = sheet[5:]
            part = part.replace("$", "").strip()
            try:
                c1, r1, c2, r2 = range_boundaries(part)
            except Exception:
                # Some rows wrap the range in prose: "Class column B4:B1015",
                # "Entire Dataset A1:AU1984". Pull the range back out.
                m2 = re.search(r"\b([A-Z]{1,3}\d+:[A-Z]{1,3}\d+)\b", part)
                if not m2:
                    continue
                try:
                    c1, r1, c2, r2 = range_boundaries(m2.group(1))
                except Exception:
                    continue
            if None in (c1, r1, c2, r2):
                continue
            targets.append((file_hint, sheet or None, (r1, c1, r2, c2)))
    return targets


def _norm(v):
    """Normalize a cell for comparison.

    Numbers and numeric-looking text must compare equal: references in this
    dataset sometimes store a value as text ("66.0") where a correct answer
    writes the number (66). Comparing those as raw strings fails a right answer
    on storage type alone, which no grader for this benchmark intends.
    """

    if v is None:
        return ""
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, (int, float)):
        return str(int(v)) if float(v).is_integer() else f"{float(v):.6g}"
    text = str(v).strip()
    if text:
        # Only plain numerics; leave "1,234", "66%", dates, and ids-with-letters
        # alone so this cannot silently equate values the task distinguishes.
        try:
            num = float(text)
        except ValueError:
            return text
        if text.lower() in ("nan", "inf", "-inf", "infinity", "-infinity"):
            return text
        return str(int(num)) if num.is_integer() else f"{num:.6g}"
    return text


def resolve_sheet(names: list[str], want: str, ordered: list[str] | None = None) -> str | None:
    """Map an ``eval_area`` sheet name onto a sheet actually in the workbook.

    ``eval_area`` names a sheet that is absent from the *reference* in 23 of 261
    tasks, for reasons that are all mechanical rather than semantic:

    * Excel truncates sheet names at 31 characters, while ``eval_area`` carries
      the untruncated name ("Comparison of Returns at Different Stages").
    * Spacing differs ("20sStatistics" vs "20s Statistics").
    * The workbook has a single sheet under a different label ("Account
      Summary" where eval_area says "Sheet1").
    * Translation drift ("1Month".."10Month" vs "January".."October").

    ``ordered`` is the full list of sheets ``eval_area`` names for this file. If
    its length matches the workbook's, positions are mapped in order, which
    resolves the drift case without guessing at meaning.
    """

    if not names:
        return None

    def key(s: str) -> str:
        return re.sub(r"\s+", "", s).strip().lower()

    by_key = {key(n): n for n in names}
    if want in names:
        return want
    if key(want) in by_key:
        return by_key[key(want)]
    # Excel caps sheet names at 31 characters of the RAW name, so truncate
    # before normalizing -- stripping spaces first shifts which characters
    # survive and the comparison misses.
    if key(want[:31]) in by_key:
        return by_key[key(want[:31])]
    for n in names:
        if key(n) == key(want[: len(n)]) or key(n[:31]) == key(want[:31]):
            return n
    if len(names) == 1:
        return names[0]
    if ordered and len(ordered) == len(names):
        try:
            return names[ordered.index(want)]
        except ValueError:
            pass
    return None


def _load_grid(path: Path, sheet: str | None = None, *, strict_sheet: bool = True,
               ordered_sheets: list[str] | None = None) -> list[list]:
    """Read a tabular output into a 2D grid, honouring an explicit sheet name.

    8% of tasks emit .csv or .xls rather than .xlsx, so grading only through
    openpyxl silently fails those on format alone.

    ``strict_sheet=False`` falls back to the first worksheet when the named one
    is absent. Use it for the model's output: 67 of 261 tasks have an
    ``eval_area`` naming a sheet the question never mentions, so requiring the
    model to guess that name would fail a correct answer on a label.
    """

    suffix = path.suffix.lower()
    if suffix in (".csv", ".txt", ".tsv"):
        import csv

        delim = "\t" if suffix == ".tsv" else ","
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                with path.open(encoding=enc, newline="") as fh:
                    return [list(r) for r in csv.reader(fh, delimiter=delim)]
            except UnicodeDecodeError:
                continue
        raise ValueError("could not decode as text")
    if suffix == ".xls":
        import pandas as pd

        try:
            df = pd.read_excel(path, sheet_name=sheet or 0, header=None, dtype=object)
        except Exception:
            if strict_sheet:
                raise
            df = pd.read_excel(path, sheet_name=0, header=None, dtype=object)
        return df.where(pd.notna(df), None).values.tolist()
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet:
        found = resolve_sheet(wb.sheetnames, sheet, ordered_sheets)
        if found is None:
            if strict_sheet:
                raise KeyError(f"sheet {sheet!r} not in {wb.sheetnames}")
            # The model could not know this name; compare positionally instead.
            found = wb.sheetnames[0]
        ws = wb[found]
    else:
        ws = wb.worksheets[0]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _at(grid: list[list], row: int, col: int):
    """1-indexed cell access, empty outside the grid."""

    if 1 <= row <= len(grid):
        r = grid[row - 1]
        if 1 <= col <= len(r):
            return r[col - 1]
    return None


def grade(
    produced: Path,
    reference: Path,
    eval_area: str,
    work: Path | None = None,
    ref_dir: Path | None = None,
) -> tuple[bool, int, int, str | None]:
    """Cell-exact comparison over every range named in eval_area.

    A target may name its own file (``[other.xlsx]#[...]``); those resolve
    against ``work`` and ``ref_dir`` so multi-output tasks grade correctly.
    """

    if produced.suffix.lower() in (".docx", ".doc", ".png", ".jpg"):
        return False, 0, 0, f"unsupported output type {produced.suffix} (not graded)"

    targets = parse_eval_area(eval_area)
    if not targets:
        return False, 0, 0, f"unparseable eval_area {eval_area[:60]!r}"

    cache: dict[tuple[str, str | None, bool], list[list]] = {}

    # All sheets eval_area names, in order, so a name-for-name drift between
    # eval_area and the workbook can be resolved positionally.
    ordered = [s for _f, s, _b in targets if s]
    seen_order: list[str] = []
    for s in ordered:
        if s not in seen_order:
            seen_order.append(s)

    def grid(path: Path, sheet: str | None, *, strict: bool) -> list[list]:
        key = (str(path), sheet, strict)
        if key not in cache:
            cache[key] = _load_grid(path, sheet, strict_sheet=strict, ordered_sheets=seen_order)
        return cache[key]

    matched = total = 0
    for file_hint, sheet, (r1, c1, r2, c2) in targets:
        ppath, rpath = produced, reference
        if file_hint and work is not None and ref_dir is not None:
            cand_p, cand_r = work / file_hint, ref_dir / file_hint
            # references are suffixed _res by convention
            if not cand_r.is_file():
                stem, _, ext = file_hint.rpartition(".")
                alt = ref_dir / f"{stem}_res.{ext}"
                cand_r = alt if alt.is_file() else cand_r
            if cand_p.is_file() and cand_r.is_file():
                ppath, rpath = cand_p, cand_r
            elif not cand_p.is_file():
                total += (r2 - r1 + 1) * (c2 - c1 + 1)
                continue
        try:
            # Reference is the golden file and does carry the right names.
            # The model's output is compared positionally when it cannot know them.
            pg = grid(ppath, sheet, strict=False)
            rg = grid(rpath, sheet, strict=True)
        except Exception as exc:
            return False, matched, total, f"load: {type(exc).__name__}: {exc}"
        for row in range(r1, r2 + 1):
            for col in range(c1, c2 + 1):
                total += 1
                if _norm(_at(pg, row, col)) == _norm(_at(rg, row, col)):
                    matched += 1
    return (matched == total and total > 0), matched, total, None


PREVIEW_ROWS = 5
PREVIEW_COLS = 12


def preview_inputs(work: Path, names: list[str]) -> str:
    """Summarize each input file: sheets, dimensions, and the first few rows.

    Deliberately leak-free. fabric_rlm's own add_excel_workbook_context requires
    a ``target_position``, and AIDABench's ``eval_area`` is a grading detail the
    model is never told -- only 1 of 261 questions states its own range. Feeding
    it in would hand over the answer location.

    This shows only what a person sees on opening the file, which is what the
    model would otherwise spend a turn or two discovering. SpreadsheetBench's
    dominant failure category is insufficient inspection, so making the shape
    visible up front is the cheap half of that fix.
    """

    blocks: list[str] = []
    for name in names:
        path = work / name
        suffix = path.suffix.lower()
        # grids and dims are always set together, keyed the same way, so the
        # rendering loop below never has to know which branch produced them.
        grids: dict[str, list] = {}
        dims: dict[str, tuple[int, int]] = {}
        head = name
        try:
            if suffix in (".csv", ".txt", ".tsv"):
                import csv

                delim = "\t" if suffix == ".tsv" else ","
                with path.open(encoding="utf-8-sig", newline="", errors="replace") as fh:
                    reader = csv.reader(fh, delimiter=delim)
                    grid = [r for _, r in zip(range(PREVIEW_ROWS + 1), reader)]
                    extra = sum(1 for _ in reader)
                head = f"{name}  (delimited text)"
                grids[""] = grid
                dims[""] = (len(grid) + extra, max((len(r) for r in grid), default=0))
            elif suffix == ".xls":
                import pandas as pd

                for k, v in pd.read_excel(path, sheet_name=None, header=None, dtype=object).items():
                    full = v.where(v.notna(), None).values.tolist()
                    grids[k] = full[: PREVIEW_ROWS + 1]
                    dims[k] = (len(full), max((len(r) for r in full), default=0))
            elif suffix in (".xlsx", ".xlsm"):
                import openpyxl

                wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
                for ws in wb.worksheets:
                    dims[ws.title] = (ws.max_row or 0, ws.max_column or 0)
                    grids[ws.title] = [
                        list(r)
                        for _, r in zip(range(PREVIEW_ROWS + 1), ws.iter_rows(values_only=True))
                    ]
                wb.close()
            else:
                blocks.append(f"- {name}  ({suffix.lstrip('.') or 'unknown'} file; open it to inspect)")
                continue
        except Exception as exc:
            blocks.append(f"- {name}  (could not preview: {type(exc).__name__})")
            continue

        lines = [f"- {head}"]
        for sheet, grid in grids.items():
            nrows, ncols = dims.get(
                sheet, (len(grid), max((len(r) for r in grid), default=0))
            )
            label = f"  sheet {sheet!r}" if sheet else "  "
            lines.append(f"{label}: {nrows} rows x {ncols} cols")
            for row in grid[:PREVIEW_ROWS]:
                cells = ["" if c is None else str(c) for c in list(row)[:PREVIEW_COLS]]
                lines.append("      " + " | ".join(c[:24] for c in cells))
        blocks.append("\n".join(lines))
    return "\n".join(blocks)


TASK_TMPL = """{question}

Working directory: {work}
All input files listed below are already in that directory.

## Input files (first {n} rows of each, for orientation)

{preview}

Row counts above are the full file; only the first rows are shown. Read the
files yourself for anything beyond this.

Write your result to a file named exactly {output_file!r} in the same directory.
Do not rename it. Do not write it anywhere else.

When finished, call SUBMIT(answer="done")."""



def acquire_run_lock(run_root: Path) -> Path | None:
    """Refuse to start if another process is already writing this run.

    Two processes appending to the same results.jsonl interleave mid-line and
    every affected row is lost. This happened twice here, each time because a
    relaunch raced a process that had not actually exited. A stale lock from a
    crash is reported with its pid so it can be cleared deliberately.
    """

    lock = run_root / ".run.lock"
    if lock.exists():
        try:
            owner = lock.read_text(encoding="utf-8").strip()
        except Exception:
            owner = "unknown"
        print(f"ERROR: {run_root.name} is locked by pid {owner}.", file=sys.stderr)
        print("  Another run is writing here. Wait for it, or if it is dead:", file=sys.stderr)
        print(f"    rm {lock}", file=sys.stderr)
        return None
    run_root.mkdir(parents=True, exist_ok=True)
    lock.write_text(str(os.getpid()), encoding="utf-8")
    return lock


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=20)
    ap.add_argument("--model", default="openrouter/minimax/minimax-m3")
    ap.add_argument("--skill", default="auto",
                    help="skill name, or 'auto' to pick per task from file types")
    ap.add_argument("--max-turns", type=int, default=14)
    ap.add_argument("--timeout", type=float, default=300.0)
    ap.add_argument("--max-tokens", type=int, default=16000)
    ap.add_argument("--temperature", type=float, default=1.0)
    ap.add_argument("--provider-order", default="minimax/fp8")
    ap.add_argument("--offset", type=int, default=0)
    ap.add_argument("--run-name", default="aida-m3")
    ap.add_argument("--only-multifile", action="store_true")
    ap.add_argument("--resume", action="store_true",
                    help="append to an existing run, skipping tasks already done")
    ap.add_argument("--regrade", default="", metavar="RUN",
                    help="re-score an existing run's saved outputs; no model calls")
    ap.add_argument("--ids", default="", help="comma-separated task ids to run")
    ap.add_argument("--no-preview", dest="preview", action="store_false",
                    help="omit the input-file preview (A/B control arm)")
    args = ap.parse_args()

    if args.regrade:
        return regrade(args.regrade)

    if not os.environ.get("OPENROUTER_API_KEY"):
        print("OPENROUTER_API_KEY not set", file=sys.stderr)
        return 2

    rows = [json.loads(l) for l in (SPLIT / "file_generation_en.jsonl").open(encoding="utf-8")]
    by_id = {r["id"]: r for r in rows}

    # Only tasks whose inputs and reference landed AND whose eval_area can be
    # cell-graded. Exclusions are printed, never silent -- a quietly shrunk
    # denominator makes a partial run look like a complete one.
    ready: list[dict] = []
    skipped = {"missing_files": [], "ungradable_area": [], "non_tabular_output": []}
    for r in rows:
        idir = SPLIT / "input" / r["id"]
        rdir = SPLIT / "reference" / r["id"]
        names = split_files(r["input_file"])
        refs = split_files(r["reference_file"])
        if (
            not idir.is_dir()
            or not all((idir / n).is_file() for n in names)
            or not any((rdir / x).is_file() for x in refs)
        ):
            skipped["missing_files"].append(r["id"])
            continue
        if not parse_eval_area(r["eval_area"]):
            skipped["ungradable_area"].append(r["id"])
            continue
        if all(Path(o).suffix.lower() in (".docx", ".doc", ".png", ".jpg")
               for o in split_files(r["output_file"])):
            skipped["non_tabular_output"].append(r["id"])
            continue
        if args.only_multifile and len(names) < 2:
            continue
        ready.append(r)

    print(f"{len(ready)} of {len(rows)} tasks runnable")
    for reason, ids in skipped.items():
        if ids:
            print(f"  excluded ({reason}): {len(ids)} -> {', '.join(ids[:12])}"
                  + (" ..." if len(ids) > 12 else ""))
    if args.ids:
        want = {i.strip() for i in args.ids.split(",") if i.strip()}
        ready = [r for r in ready if r["id"] in want]
    todo = ready[args.offset : args.offset + args.limit]
    if not todo:
        print("nothing to run")
        return 1

    run_root = BENCH / "runs" / args.run_name
    work_root = run_root / "work"
    for p in (run_root, work_root):
        p.mkdir(parents=True, exist_ok=True)
    lock = acquire_run_lock(run_root)
    if lock is None:
        return 2
    results_path = run_root / "results.jsonl"
    done: set[str] = set()
    if args.resume and results_path.is_file():
        for line in results_path.open(encoding="utf-8"):
            try:
                done.add(json.loads(line)["id"])
            except Exception:
                continue
        todo = [r for r in todo if r["id"] not in done]
        print(f"resuming: {len(done)} already done, {len(todo)} remaining")

    lm = dspy.LM(
        args.model,
        api_key=os.environ["OPENROUTER_API_KEY"],
        api_base="https://openrouter.ai/api/v1",
        max_tokens=args.max_tokens,
        temperature=args.temperature,
        extra_body={"usage": {"include": True}}
        | (
            {"provider": {"order": args.provider_order.split(","), "allow_fallbacks": False}}
            if args.provider_order
            else {}
        ),
    )
    # Custom dirs now layer over the packaged skills, so point at the local
    # dir directly -- no merging, and excel_modify stays available.
    local_skills = ROOT / "_local" / "skills"
    loader = SkillLoader(skill_dir=str(local_skills)) if local_skills.is_dir() else SkillLoader()
    n_pass = 0

    if not todo:
        print("nothing left to run")
        return 0

    with results_path.open("a" if args.resume else "w", encoding="utf-8") as out:
        for i, rec in enumerate(todo, 1):
            try:
                _run_one(rec, i, len(todo), args, lm, loader, work_root, out)
            except Exception as exc:
                # A benchmark must never die on one task. Three runs were lost
                # this way -- a console encoding error, an unbound local in the
                # preview builder -- each after an hour of work. Record it and
                # keep going.
                traceback.print_exc(limit=3)
                out.write(json.dumps({
                    "id": rec["id"], "passed": False, "cells_matched": 0, "cells_total": 0,
                    "grade_err": None, "run_err": f"harness: {type(exc).__name__}: {exc}"[:300],
                    "n_inputs": len(split_files(rec["input_file"])), "n_turns": 0,
                    "prompt_tokens": 0, "completion_tokens": 0, "elapsed_seconds": 0.0,
                    "model": args.model, "skill": "",
                }) + "\n")
                out.flush()
                print(f"  [{i}/{len(todo)}] {rec['id']:<8} HARNESS ERROR {type(exc).__name__}")
                continue
            n_pass += 1 if _LAST_PASS[0] else 0

    print(f"\n{n_pass}/{len(todo)} = {100*n_pass/len(todo):.1f}%  ->  {results_path}")
    return 0


# Set by _run_one so the caller can tally passes without re-reading the file.
_LAST_PASS = [False]


def _run_one(rec, i, n_total, args, lm, loader, work_root, out):
    """Execute and grade one task. Raising here costs one task, not the run."""

    if True:
        if True:
            qid = rec["id"]
            names = split_inputs(rec["input_file"])
            work = work_root / qid
            # ignore_errors: on Windows a stale handle from another process
            # makes rmtree raise, failing the task for a non-model reason.
            if work.exists():
                shutil.rmtree(work, ignore_errors=True)
            work.mkdir(parents=True, exist_ok=True)
            for n in names:
                shutil.copy2(SPLIT / "input" / qid / n, work / n)

            # Route by the file types actually involved. Handing a docx or PDF
            # task the Excel playbook is worse than handing it nothing.
            exts = {Path(n).suffix.lower() for n in names} | {Path(rec["output_file"]).suffix.lower()}
            if args.skill != "auto":
                skills = [args.skill]
            else:
                skills = ["excel_modify"]
                if exts & {".docx", ".doc"}:
                    skills = ["docx_documents"] + ([] if exts <= {".docx", ".doc"} else ["excel_modify"])
                if ".pdf" in exts:
                    skills = ["pdf_document_analysis"] + [s for s in skills if s != "pdf_document_analysis"]

            task = TASK_TMPL.format(
                question=rec["question"].strip(),
                work=str(work),
                n=PREVIEW_ROWS,
                preview=preview_inputs(work, names) if args.preview else "(not provided)",
                output_file=", ".join(repr(o) for o in split_files(rec["output_file"])),
            )
            inputs = {
                re.sub(r"\W+", "_", Path(n).stem)[:40] or f"file{j}": File(str(work / n))
                for j, n in enumerate(names)
            }

            t0 = time.perf_counter()
            err = None
            try:
                rlm = RLM.from_task(
                    task=task,
                    inputs=inputs,
                    outputs=["answer"],
                    lm=lm,
                    skill_loader=loader,
                    skills=skills,
                    max_turns=args.max_turns,
                    timeout=args.timeout,
                )
                result = rlm.run()
                save_trace(
                    work_root.parent / "traces" / f"{qid}.json",
                    task_id=qid,
                    task_text=task,
                    result=result,
                    config={
                        "model": args.model,
                        "skill": ",".join(skills),
                        "max_turns": args.max_turns,
                        "temperature": args.temperature,
                        "split": "file_generation_en",
                    },
                )
                turns = len(result.trajectory.turns)
                ptok = result.total_prompt_tokens or 0
                ctok = result.total_completion_tokens or 0
            except Exception as exc:
                err = f"{type(exc).__name__}: {exc}"
                traceback.print_exc(limit=2)
                turns = ptok = ctok = 0
            elapsed = time.perf_counter() - t0

            outs = split_files(rec["output_file"])
            refs = split_files(rec["reference_file"])
            rdir = SPLIT / "reference" / qid
            passed, matched, total, gerr = True, 0, 0, None
            for oi, o in enumerate(outs):
                produced = work / o
                ref = rdir / (refs[oi] if oi < len(refs) else refs[0])
                if not ref.is_file():
                    continue
                if not produced.is_file():
                    passed = False
                    gerr = gerr or f"output file not created: {o}"
                    continue
                ok, m, t, e = grade(produced, ref, rec["eval_area"], work, rdir)
                passed = passed and ok
                matched += m
                total += t
                gerr = gerr or e
            if total == 0 and gerr is None:
                passed, gerr = False, "no gradable output/reference pair"

            row = {
                "id": qid,
                "passed": bool(passed),
                "cells_matched": matched,
                "cells_total": total,
                "grade_err": gerr,
                "run_err": err,
                "n_inputs": len(names),
                "n_turns": turns,
                "prompt_tokens": ptok,
                "completion_tokens": ctok,
                "elapsed_seconds": round(elapsed, 2),
                "model": args.model,
                "skill": ",".join(skills),
            }
            out.write(json.dumps(row) + "\n")
            out.flush()
            flag = "PASS" if passed else "fail"
            print(
                f"  [{i}/{n_total}] {qid:<8} {flag}  "
                f"{matched}/{total} cells  {len(names)} input(s)  "
                f"{turns} turns  {elapsed:.0f}s"
                + (f"  ({gerr})" if gerr else "")
            )
            _LAST_PASS[0] = bool(passed)
    return None


def regrade(run_name: str) -> int:
    """Re-score a finished run from its saved work dirs. No model calls.

    Lets a grader fix apply to results already collected, so every task is
    scored by the same rules rather than by whichever version was live when it
    happened to run.
    """

    run_root = BENCH / "runs" / run_name
    results = run_root / "results.jsonl"
    if not results.is_file():
        print(f"no results at {results}", file=sys.stderr)
        return 1
    rows = [json.loads(line) for line in results.open(encoding="utf-8")]
    meta = {
        str(r["id"]): r
        for r in (
            json.loads(line)
            for line in (SPLIT / "file_generation_en.jsonl").open(encoding="utf-8")
        )
    }
    before = sum(1 for r in rows if r["passed"])
    changed = 0
    for row in rows:
        rec = meta[row["id"]]
        work = run_root / "work" / row["id"]
        produced = work / rec["output_file"]
        ref = SPLIT / "reference" / row["id"] / rec["reference_file"]
        if produced.is_file():
            passed, matched, total, gerr = grade(
                produced, ref, rec["eval_area"], work, SPLIT / "reference" / row["id"]
            )
        else:
            passed, matched, total, gerr = False, 0, 0, "output file not created"
        if bool(passed) != bool(row["passed"]):
            changed += 1
        row.update(
            passed=bool(passed), cells_matched=matched, cells_total=total, grade_err=gerr
        )
    out = run_root / "results_regraded.jsonl"
    with out.open("w", encoding="utf-8") as fh:
        for row in rows:
            fh.write(json.dumps(row) + "\n")
    after = sum(1 for r in rows if r["passed"])
    print(f"regraded {len(rows)} tasks from {run_name}")
    print(f"  before {before}/{len(rows)} = {100*before/len(rows):.1f}%")
    print(f"  after  {after}/{len(rows)} = {100*after/len(rows):.1f}%   ({changed} verdicts changed)")
    print(f"  -> {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
